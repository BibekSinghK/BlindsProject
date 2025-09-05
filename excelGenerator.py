from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import os

class ExcelGenerator:
    def __init__(self, person):
        self.person = person
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = f"{self.person.name} Cost Sheet"

        self.center = Alignment(horizontal="center", vertical="center")
        self.boldCenter = Font(bold=True)
        self.two_decimal = '0.00'
    
    def write_blinds(self):
        
        labels = ["Customer", "Item", "Location", "Fabric", "No", "Width", "Height",
          "Width (mm)", "Height (mm)", "Qty/m^2", "One set price", "Set", "Total Price",
          "Control/Chain/Plastic Chain", "Bracket"]
        
        column = 1 
        for label in labels:
            cell = self.worksheet.cell(row=1, column=column, value=label)
            cell.alignment = self.center
            cell.font = self.boldCenter
            column += 1
        
        row = 1
        for blind in self.person.blinds:
            row += 1 

            self.worksheet.cell(row=row, column=1, value=self.person.name).alignment = self.center
            self.worksheet.cell(row=row, column=2, value=blind.blind_type).alignment = self.center
            self.worksheet.cell(row=row, column=3, value=blind.location).alignment = self.center
            self.worksheet.cell(row=row, column=4, value=blind.fabric).alignment = self.center
            self.worksheet.cell(row=row, column=5, value=blind.blindNo).alignment = self.center
            self.worksheet.cell(row=row, column=6, value=float(blind.width)).alignment = self.center
            self.worksheet.cell(row=row, column=7, value=float(blind.height)).alignment = self.center

            self.worksheet.cell(row=row, column=8).value = f"=F{row} * 25.4"
            self.worksheet.cell(row=row, column=9).value = f"=G{row} * 25.4"

            if blind.blind_type == 'Roller':
                self.worksheet.cell(row=row, column=10).value = f"=((I{row} + 200) * H{row}) / 1000000"
                self.worksheet.cell(row=row, column=11).value = f"=J{row} * {blind.price} + (H{row}/1000*7) + (H{row}/1000*1.9)"
            else:
                self.worksheet.cell(row=row, column=10).value = f"=((I{row} + 150) * H{row}) / 1000000"
                self.worksheet.cell(row=row, column=11).value = f"=J{row} * {blind.price}"

            self.worksheet.cell(row=row, column=12, value=int(blind.quantity)).alignment = self.center
            self.worksheet.cell(row=row, column=13).value = f"=K{row} * {blind.quantity}"

            self.worksheet.cell(row=row, column=14, value=f"{blind.controlPos} {blind.controlMat} {blind.control}")
            self.worksheet.cell(row=row, column=15, value=blind.bracket)

        sub_row = self.person.blindCount + 3
        ship_row = self.person.blindCount + 4
        total_row = self.person.blindCount + 6

        self.worksheet.cell(row=sub_row, column=12, value="Subtotal:").font = self.boldCenter
        self.worksheet.cell(row=sub_row, column=13).value = f"=SUM(M2:M{self.person.blindCount + 1})"

        self.worksheet.cell(row=ship_row, column=12, value="Shipping:").font = self.boldCenter
        self.worksheet.cell(row=ship_row, column=13).value = f"=SUM(L2:L{self.person.blindCount + 1}) * 33"

        self.worksheet.cell(row=total_row, column=12, value="Total Price:").font = self.boldCenter
        self.worksheet.cell(row=total_row, column=13).value = f"=M{sub_row} + M{ship_row}"

        for col in range(1, 16):
            column_letter = get_column_letter(col)
            self.worksheet.column_dimensions[column_letter].auto_size = True 

        output_dir = "cost_sheets"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filename = f"{self.person.name}_cost_sheet.xlsx"
        file_path = os.path.join(output_dir, filename)
        self.workbook.save(file_path)